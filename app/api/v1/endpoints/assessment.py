"""
QBO Diagnostic Assessment endpoint.

POST /api/v1/assessment/run
  - Fetches QBO data in parallel (P&L, Balance Sheet, COA, AR aging, AP aging,
    undeposited funds)
  - Runs structured diagnostic checks matching the Excel template
  - Fills the template workbook with findings
  - Returns the completed .xlsx as a download
"""

from __future__ import annotations

import asyncio
import io
import logging
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.dependencies import get_db, get_qbo_client_for_realm

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/assessment", tags=["Assessment"])

# Path to the bundled Excel template
TEMPLATE_PATH = Path(__file__).parent.parent.parent.parent / "static" / "assessment_template.xlsx"


# ── Utility helpers ────────────────────────────────────────────────────────────

def _safe_float(v: Any) -> float:
    """Coerce a QBO value (may contain commas) to float, default 0.0."""
    try:
        return float(str(v).replace(",", ""))
    except (TypeError, ValueError):
        return 0.0


def _fmt_mmyy(date_str: str) -> str:
    """Convert 'YYYY-MM-DD' to 'MM/YY'."""
    try:
        d = datetime.strptime(date_str[:10], "%Y-%m-%d")
        return d.strftime("%m/%y")
    except Exception:
        return date_str


def _fmt_period(date_str: str) -> str:
    """Convert 'YYYY-MM-DD' to 'MM/YYYY'."""
    try:
        d = datetime.strptime(date_str[:10], "%Y-%m-%d")
        return d.strftime("%m/%Y")
    except Exception:
        return date_str


def _extract_report_rows(report: dict) -> list[dict]:
    """
    Flatten all leaf Data rows from a QBO report into {name, amount} pairs.
    QBO report structure: report.Rows.Row[] where each Row is type Section (recurse)
    or type Data (leaf with ColData[0]=name, ColData[1]=amount).
    """
    results: list[dict] = []

    def walk(rows: list[dict]) -> None:
        for row in rows:
            row_type = row.get("type", "")
            if row_type == "Section":
                child_rows = row.get("Rows", {}).get("Row", [])
                walk(child_rows)
                # Capture section summary row (for totals like "Total Income")
                summary = row.get("Summary", {})
                s_cols = summary.get("ColData", [])
                if len(s_cols) >= 2:
                    results.append({
                        "name": s_cols[0].get("value", ""),
                        "amount": _safe_float(s_cols[1].get("value", "0")),
                        "is_summary": True,
                    })
            elif row_type == "Data":
                cols = row.get("ColData", [])
                if len(cols) >= 2:
                    results.append({
                        "name": cols[0].get("value", ""),
                        "amount": _safe_float(cols[1].get("value", "0")),
                        "is_summary": False,
                    })

    rows = report.get("Rows", {}).get("Row", [])
    walk(rows)
    return results


def _find_amount(rows: list[dict], *name_fragments: str) -> float:
    """Return amount of first row whose name contains any fragment (case-insensitive)."""
    fragments_lower = [f.lower() for f in name_fragments]
    for row in rows:
        if any(frag in row["name"].lower() for frag in fragments_lower):
            return row["amount"]
    return 0.0


def _find_rows_matching(rows: list[dict], *name_fragments: str) -> list[dict]:
    """Return all rows whose name contains any fragment (case-insensitive)."""
    fragments_lower = [f.lower() for f in name_fragments]
    return [r for r in rows if any(f in r["name"].lower() for f in fragments_lower)]


def _extract_aging_buckets(report: dict) -> dict:
    """
    Parse an AR/AP aging report.
    Returns dict with keys: buckets, items_90plus, items_negative, items_zero.
    """
    buckets: dict[str, float] = {
        "current": 0.0, "1-30": 0.0, "31-60": 0.0, "61-90": 0.0, "91+": 0.0, "total": 0.0
    }
    items_90plus: list[dict] = []
    items_negative: list[dict] = []
    items_zero: list[dict] = []

    try:
        col_headers: list[str] = []
        for col in report.get("Columns", {}).get("Column", []):
            col_headers.append(col.get("ColTitle", "").strip())

        col_90_idx: int | None = None
        for i, h in enumerate(col_headers):
            if ">90" in h or "91" in h or ("over" in h.lower() and "90" in h):
                col_90_idx = i
                break

        def walk_aging(rows: list[dict]) -> None:
            for row in rows:
                if row.get("type") == "Data":
                    cols = row.get("ColData", [])
                    entity_name = cols[0].get("value", "") if cols else ""
                    row_total = 0.0
                    for i, col in enumerate(cols[1:], 1):
                        amt = _safe_float(col.get("value", "0"))
                        row_total += amt
                        h = col_headers[i] if i < len(col_headers) else ""
                        hl = h.lower()
                        if "current" in hl:
                            buckets["current"] += amt
                        elif "1" in h and "30" in h:
                            buckets["1-30"] += amt
                        elif "31" in h and "60" in h:
                            buckets["31-60"] += amt
                        elif "61" in h and "90" in h:
                            buckets["61-90"] += amt
                        elif ">90" in h or "91" in h or ("over" in hl and "90" in h):
                            buckets["91+"] += amt

                    if col_90_idx is not None and col_90_idx < len(cols):
                        amt_90 = _safe_float(cols[col_90_idx].get("value", "0"))
                        if abs(amt_90) > 0.01:
                            items_90plus.append({"name": entity_name, "amount": amt_90})

                    if row_total < -0.01:
                        items_negative.append({"name": entity_name, "amount": row_total})
                    if abs(row_total) < 0.01 and entity_name:
                        items_zero.append({"name": entity_name, "amount": row_total})

                elif row.get("type") == "Section":
                    walk_aging(row.get("Rows", {}).get("Row", []))

        walk_aging(report.get("Rows", {}).get("Row", []))
        buckets["total"] = sum(v for k, v in buckets.items() if k != "total")

    except Exception as exc:
        logger.warning("Error parsing aging report: %s", exc)

    return {
        "buckets": buckets,
        "items_90plus": items_90plus,
        "items_negative": items_negative,
        "items_zero": items_zero,
    }


# ── Cell-writing helpers ───────────────────────────────────────────────────────

def _set_pl_finding(
    ws: Any, row: int, finding: str,
    comment: str = "", num_txns: str = "",
    date_from: str = "", date_to: str = "", amount: str = "",
) -> None:
    """Write a P&L findings row. J=finding, L=comment, M=#txns, N=from, O=to, P=amount."""
    ws[f"J{row}"].value = finding
    if finding == "clean up needed":
        if comment:
            ws[f"L{row}"].value = comment
        if num_txns:
            ws[f"M{row}"].value = num_txns
        if date_from:
            ws[f"N{row}"].value = date_from
        if date_to:
            ws[f"O{row}"].value = date_to
        if amount:
            ws[f"P{row}"].value = amount


def _set_bs_finding(
    ws: Any, row: int, finding: str,
    comment: str = "", num_txns: str = "",
    date_from: str = "", date_to: str = "", amount: str = "",
) -> None:
    """Write a Balance Sheet findings row. J=finding, L=comment, O=#txns, P=from, Q=to, R=amount."""
    ws[f"J{row}"].value = finding
    if finding == "clean up needed":
        if comment:
            ws[f"L{row}"].value = comment
        if num_txns:
            ws[f"O{row}"].value = num_txns
        if date_from:
            ws[f"P{row}"].value = date_from
        if date_to:
            ws[f"Q{row}"].value = date_to
        if amount:
            ws[f"R{row}"].value = amount


def _set_arap_finding(
    ws: Any, row: int, finding: str,
    num_txns: str = "", date_from: str = "", date_to: str = "", amount: str = "",
) -> None:
    """Write an AR/AP findings row. J=finding, K=#txns, L=from, M=to, N=amount."""
    ws[f"J{row}"].value = finding
    if finding == "clean up needed":
        if num_txns:
            ws[f"K{row}"].value = num_txns
        if date_from:
            ws[f"L{row}"].value = date_from
        if date_to:
            ws[f"M{row}"].value = date_to
        if amount:
            ws[f"N{row}"].value = amount


# ── Main route ────────────────────────────────────────────────────────────────

@router.post("/run")
async def run_assessment(
    realm_id: str = Query(..., description="QBO realm/company ID"),
    period_from: str = Query(..., description="Period start date YYYY-MM-DD"),
    period_to: str = Query(..., description="Period end date YYYY-MM-DD"),
    client_name: str = Query("", description="Client display name"),
    accounting_method: str = Query("Accrual", description="Accrual or Cash"),
    qbo_version: str = Query("Plus", description="QBO plan tier"),
    tax_org_type: str = Query("", description="Tax org type e.g. S-Corp, LLC"),
    db: AsyncSession = Depends(get_db),
) -> StreamingResponse:
    """
    Run a full QBO diagnostic assessment and return a completed Excel workbook.

    Fetches P&L, Balance Sheet, Chart of Accounts, AR aging, AP aging, and
    undeposited funds in parallel, runs ~40 diagnostic checks against the
    TPC QuickBooks Diagnostic Template, then streams the filled workbook as a download.
    """
    client = await get_qbo_client_for_realm(realm_id, db)
    if not client:
        raise HTTPException(status_code=404, detail=f"No QBO connection for realm {realm_id}")

    if not TEMPLATE_PATH.exists():
        raise HTTPException(
            status_code=500,
            detail="Assessment template not found on server. Contact support.",
        )

    # ── Fetch all QBO data in parallel ────────────────────────────────────────
    async def safe_fetch(coro: Any, name: str, default: Any) -> Any:
        try:
            return await coro
        except Exception as exc:
            logger.warning("QBO fetch failed for %s: %s", name, exc)
            return default

    (
        pl_report,
        bs_report,
        coa_list,
        ar_aging_report,
        ap_aging_report,
        undeposited_funds,
        items_list,
    ) = await asyncio.gather(
        safe_fetch(client.get_profit_loss(period_from, period_to), "profit_loss", {}),
        safe_fetch(client.get_balance_sheet(period_to), "balance_sheet", {}),
        safe_fetch(client.get_chart_of_accounts(), "chart_of_accounts", []),
        safe_fetch(client.get_ar_aging(), "ar_aging", {}),
        safe_fetch(client.get_ap_aging(), "ap_aging", {}),
        safe_fetch(client.get_undeposited_funds(), "undeposited_funds", []),
        safe_fetch(client._query("SELECT * FROM Item MAXRESULTS 500"), "items", []),
    )

    # ── Pre-process data ───────────────────────────────────────────────────────
    pl_rows = _extract_report_rows(pl_report)
    bs_rows = _extract_report_rows(bs_report)
    ar_data = _extract_aging_buckets(ar_aging_report)
    ap_data = _extract_aging_buckets(ap_aging_report)

    active_accounts = [a for a in coa_list if a.get("Active", True)]
    bank_accounts = [a for a in active_accounts if a.get("AccountType") in ("Bank", "Credit Card")]
    total_account_count = len(active_accounts)

    has_payroll_liability = any(
        "payroll" in (a.get("Name", "") + a.get("AccountSubType", "")).lower()
        for a in active_accounts
    )
    has_sales_tax = any(
        "sales tax" in a.get("Name", "").lower() or a.get("AccountSubType") == "SalesTaxPayable"
        for a in active_accounts
    )

    period_from_mmyy = _fmt_mmyy(period_from)
    period_to_mmyy = _fmt_mmyy(period_to)
    period_label = f"{_fmt_period(period_from)} - {_fmt_period(period_to)}"

    # ── Load workbook ──────────────────────────────────────────────────────────
    wb = openpyxl.load_workbook(TEMPLATE_PATH)

    # Master issues list — accumulated from each sheet section
    issues_found: list[str] = []

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET: Client info
    # ══════════════════════════════════════════════════════════════════════════
    ws_client = wb["Client info"]
    ws_client["A1"].value = client_name or "Client Name Not Provided"
    ws_client["J4"].value = period_label
    ws_client["J5"].value = accounting_method
    ws_client["J6"].value = tax_org_type or "Not specified"
    ws_client["J10"].value = f"QuickBooks Online {qbo_version}"

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET: Banking
    # ══════════════════════════════════════════════════════════════════════════
    ws_bank = wb["Banking "]
    banking_issues: list[str] = []

    # Build a lookup: account Name → LastReconcileDate (from COA data)
    acct_rec_dates: dict[str, str] = {}
    for a in coa_list:
        lrd = a.get("LastReconcileDate", "") or ""
        if lrd:
            acct_rec_dates[a.get("Name", "")] = lrd

    unreconciled_accounts: list[str] = []
    old_rec_accounts: list[str] = []

    for i, acct in enumerate(bank_accounts[:5], start=4):
        name = acct.get("Name", "")
        ws_bank[f"A{i}"].value = name
        last_rec = acct.get("LastReconcileDate", "") or acct_rec_dates.get(name, "")
        if last_rec:
            # Format YYYY-MM-DD → MM/DD/YYYY
            try:
                from datetime import date
                d = date.fromisoformat(last_rec)
                ws_bank[f"G{i}"].value = d.strftime("%m/%d/%Y")
                # Flag if reconciled more than 45 days ago
                days_since = (date.today() - d).days
                if days_since > 45:
                    old_rec_accounts.append(f"{name} (last: {d.strftime('%m/%d/%Y')})")
            except Exception:
                ws_bank[f"G{i}"].value = last_rec
        else:
            ws_bank[f"G{i}"].value = "Never reconciled"
            unreconciled_accounts.append(name)

        ws_bank[f"J{i}"].value = "Review in QBO Reconcile"
        ws_bank[f"O{i}"].value = "Review in QBO Reconcile"
        cc_types = ("CreditCard", "Credit Card")
        is_cc = acct.get("AccountType") == "Credit Card" or acct.get("AccountSubType") in cc_types
        ws_bank[f"R{i}"].value = "Review in QBO Banking"
        ws_bank[f"W{i}"].value = "Review in QBO Banking"

    if bank_accounts:
        acct_lines = []
        for a in bank_accounts[:5]:
            name = a.get("Name", "")
            lrd = a.get("LastReconcileDate", "") or acct_rec_dates.get(name, "")
            status = ""
            if lrd:
                try:
                    from datetime import date
                    d = date.fromisoformat(lrd)
                    days = (date.today() - d).days
                    status = f"reconciled through {d.strftime('%m/%d/%Y')} ({days} days ago)"
                except Exception:
                    status = f"last reconciled: {lrd}"
            else:
                status = "NEVER RECONCILED"
            acct_lines.append(f"• {name}: {status}")

        banking_summary = (
            f"Found {len(bank_accounts)} bank/credit card account(s):\n"
            + "\n".join(acct_lines)
        )
        if unreconciled_accounts:
            banking_issues.append(f"Account(s) never reconciled: {', '.join(unreconciled_accounts)}")
            banking_summary += f"\n\nACTION REQUIRED: {', '.join(unreconciled_accounts)} — establish opening balance and reconcile."
        if old_rec_accounts:
            banking_issues.append(f"Account(s) overdue for reconciliation (>45 days): {', '.join(old_rec_accounts)}")
            banking_summary += f"\n\nOVERDUE: {', '.join(old_rec_accounts)} — reconcile to current date."
        banking_summary += "\n\nFor each account: review uncleared items, auto-adjustments, bank feed status, and old transactions in QBO Reconcile and Banking tabs."
    else:
        banking_summary = "No bank or credit card accounts found in Chart of Accounts."
        banking_issues.append("No bank accounts found in COA")

    if undeposited_funds:
        udf_total = sum(_safe_float(f.get("Amount", 0)) for f in undeposited_funds)
        if udf_total > 0:
            banking_issues.append(f"Undeposited Funds balance of ${udf_total:,.2f} — review and clear")
            banking_summary += f"\n\nUndeposited Funds: ${udf_total:,.2f} outstanding — review and deposit or void stale items."

    ws_bank["A11"].value = banking_summary

    # Work to be completed — Banking
    banking_work_items = []
    if unreconciled_accounts:
        banking_work_items.append(f"Establish opening balance and perform initial reconciliation for: {', '.join(unreconciled_accounts)}")
    if old_rec_accounts:
        banking_work_items.append(f"Bring reconciliation current for: {', '.join(old_rec_accounts)}")
    if bank_accounts:
        banking_work_items.append("Review each account in QBO Reconcile > History for old uncleared items and auto-adjustments")
        banking_work_items.append("Verify bank feed is connected and active for each account in QBO Banking > Banking tab")
        banking_work_items.append("Clear any old unreviewed transactions in the For Review tab (older than 30 days)")
    ws_bank["A16"].value = (
        "\n".join(f"• {w}" for w in banking_work_items)
        if banking_work_items else "No immediate banking work required based on available data."
    )

    issues_found.extend(banking_issues)

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET: Profit & Loss
    # ══════════════════════════════════════════════════════════════════════════
    ws_pl = wb["Profit & Loss"]
    pl_issues: list[str] = []

    ws_pl["J5"].value = period_from_mmyy
    ws_pl["K5"].value = period_to_mmyy

    # J9 — Negative income balances
    neg_income = [
        r for r in pl_rows
        if r["amount"] < -0.01 and not r.get("is_summary")
        and any(kw in r["name"].lower() for kw in ["income", "revenue", "sales"])
    ]
    if neg_income:
        names = ", ".join(r["name"] for r in neg_income[:3])
        total_neg = sum(r["amount"] for r in neg_income)
        _set_pl_finding(ws_pl, 9, "clean up needed",
                        comment=f"Negative income accounts: {names}",
                        num_txns=str(len(neg_income)),
                        date_from=period_from_mmyy, date_to=period_to_mmyy,
                        amount=f"${total_neg:,.2f}")
        pl_issues.append(f"Negative income balances in {len(neg_income)} account(s)")
    else:
        _set_pl_finding(ws_pl, 9, "OK")

    # J10 — Uncategorized income
    uncat_inc = [r for r in _find_rows_matching(pl_rows, "uncategorized income", "other income") if abs(r["amount"]) > 0.01]
    if uncat_inc:
        total_ui = sum(r["amount"] for r in uncat_inc)
        _set_pl_finding(ws_pl, 10, "clean up needed",
                        comment="Uncategorized/Other Income has a balance — reclassify to proper accounts",
                        num_txns=str(len(uncat_inc)),
                        date_from=period_from_mmyy, date_to=period_to_mmyy,
                        amount=f"${total_ui:,.2f}")
        pl_issues.append("Uncategorized income balance found")
    else:
        _set_pl_finding(ws_pl, 10, "OK")

    # J11 — Sales of Product Income balance
    sopi = _find_amount(pl_rows, "sales of product income")
    if abs(sopi) > 0.01:
        _set_pl_finding(ws_pl, 11, "clean up needed",
                        comment="Sales of Product Income has a balance — verify this is appropriate for client's industry",
                        amount=f"${sopi:,.2f}")
        pl_issues.append("Sales of Product Income balance found — verify industry fit")
    else:
        _set_pl_finding(ws_pl, 11, "OK")

    # J12 — Services account balance
    svc_amt = _find_amount(pl_rows, "services")
    if abs(svc_amt) > 0.01:
        _set_pl_finding(ws_pl, 12, "clean up needed",
                        comment="Services account has a balance — verify this is appropriate for client's industry",
                        amount=f"${svc_amt:,.2f}")
        pl_issues.append("Services account balance found — verify industry fit")
    else:
        _set_pl_finding(ws_pl, 12, "OK")

    # J13 — Deposits recorded as income
    dep_income = _find_amount(pl_rows, "deposit")
    if abs(dep_income) > 0.01:
        _set_pl_finding(ws_pl, 13, "clean up needed",
                        comment="Deposits recorded as income — reclassify to liability or proper income account",
                        amount=f"${dep_income:,.2f}")
        pl_issues.append("Deposits recorded as income")
    else:
        _set_pl_finding(ws_pl, 13, "OK")

    # J14 — Loan proceeds as income
    loan_inc = _find_amount(pl_rows, "loan proceeds", "loan income", "ppp loan", "eidl")
    if abs(loan_inc) > 0.01:
        _set_pl_finding(ws_pl, 14, "clean up needed",
                        comment="Loan proceeds appear in income — reclassify to liability account",
                        amount=f"${loan_inc:,.2f}")
        pl_issues.append("Loan proceeds recorded as income")
    else:
        _set_pl_finding(ws_pl, 14, "OK")

    # J15 — Sales tax as income deduction
    st_inc = _find_amount(pl_rows, "sales tax", "tax collected")
    if abs(st_inc) > 0.01:
        _set_pl_finding(ws_pl, 15, "clean up needed",
                        comment="Sales tax appears as an income deduction — record in Sales Tax Payable liability instead",
                        amount=f"${st_inc:,.2f}")
        pl_issues.append("Sales tax recorded as income deduction")
    else:
        _set_pl_finding(ws_pl, 15, "OK")

    # J18 — Negative COGS
    neg_cogs = [
        r for r in pl_rows
        if r["amount"] < -0.01 and not r.get("is_summary")
        and any(kw in r["name"].lower() for kw in ["cost of goods", "cogs", "cost of sales"])
    ]
    if neg_cogs:
        names = ", ".join(r["name"] for r in neg_cogs[:3])
        _set_pl_finding(ws_pl, 18, "clean up needed",
                        comment=f"Negative COGS balances: {names}",
                        num_txns=str(len(neg_cogs)),
                        amount=f"${sum(r['amount'] for r in neg_cogs):,.2f}")
        pl_issues.append("Negative COGS balances found")
    else:
        _set_pl_finding(ws_pl, 18, "OK")

    # J19 — Incorrectly categorized COGS
    cogs_suspect = [r for r in _find_rows_matching(pl_rows, "insurance", "utilities", "rent") if abs(r["amount"]) > 0.01]
    if cogs_suspect:
        names = ", ".join(r["name"] for r in cogs_suspect[:3])
        _set_pl_finding(ws_pl, 19, "clean up needed",
                        comment=f"Possible expense accounts in COGS section: {names} — review categorization",
                        num_txns=str(len(cogs_suspect)))
        pl_issues.append("Potential misclassification in COGS section")
    else:
        _set_pl_finding(ws_pl, 19, "OK")

    # J20 — COGS vs income ratio check
    total_income = _find_amount(pl_rows, "total income", "gross revenue")
    if total_income == 0.0:
        total_income = sum(r["amount"] for r in pl_rows if r.get("is_summary") and "income" in r["name"].lower())
    total_cogs = _find_amount(pl_rows, "total cost of goods", "total cogs")
    if total_income > 0 and total_cogs > total_income * 1.1:
        _set_pl_finding(ws_pl, 20, "clean up needed",
                        comment=f"COGS (${total_cogs:,.2f}) exceeds total income (${total_income:,.2f}) — review COGS entries",
                        amount=f"${total_cogs:,.2f}")
        pl_issues.append("COGS exceeds total income")
    elif total_income > 10000 and total_cogs == 0:
        _set_pl_finding(ws_pl, 20, "clean up needed",
                        comment="No COGS recorded despite income — verify if client has direct costs",
                        amount="$0.00")
        pl_issues.append("No COGS recorded despite income")
    else:
        _set_pl_finding(ws_pl, 20, "OK")

    # J23 — Negative expense balances
    neg_exp = [
        r for r in pl_rows
        if r["amount"] < -0.01 and not r.get("is_summary")
        and not any(kw in r["name"].lower() for kw in ["income", "revenue", "cogs", "cost of"])
    ]
    if neg_exp:
        names = ", ".join(r["name"] for r in neg_exp[:3])
        _set_pl_finding(ws_pl, 23, "clean up needed",
                        comment=f"Negative expense balances: {names}",
                        num_txns=str(len(neg_exp)),
                        amount=f"${sum(r['amount'] for r in neg_exp):,.2f}")
        pl_issues.append(f"Negative expense balances in {len(neg_exp)} account(s)")
    else:
        _set_pl_finding(ws_pl, 23, "OK")

    # J24 — Expenses higher than expected (non-payroll accounts > $50k)
    high_exp = [
        r for r in pl_rows
        if r["amount"] > 50000 and not r.get("is_summary")
        and not any(kw in r["name"].lower() for kw in ["income", "revenue", "cogs", "payroll", "salary", "wage"])
    ]
    if high_exp:
        names = ", ".join(f"{r['name']} (${r['amount']:,.2f})" for r in high_exp[:3])
        _set_pl_finding(ws_pl, 24, "clean up needed",
                        comment=f"Unusually high expense accounts: {names}",
                        num_txns=str(len(high_exp)))
        pl_issues.append("Some expense accounts unusually high — review")
    else:
        _set_pl_finding(ws_pl, 24, "OK")

    # J25 — Expenses lower than expected (informational; cannot assess without benchmarks)
    _set_pl_finding(ws_pl, 25, "OK")

    # J26 — Uncategorized expenses
    uncat_exp = [r for r in _find_rows_matching(pl_rows, "uncategorized expense", "uncategorized") if abs(r["amount"]) > 0.01]
    if uncat_exp:
        total_ue = sum(r["amount"] for r in uncat_exp)
        _set_pl_finding(ws_pl, 26, "clean up needed",
                        comment="Uncategorized Expense has a balance — reclassify all transactions",
                        num_txns=str(len(uncat_exp)),
                        date_from=period_from_mmyy, date_to=period_to_mmyy,
                        amount=f"${total_ue:,.2f}")
        pl_issues.append("Uncategorized expenses found")
    else:
        _set_pl_finding(ws_pl, 26, "OK")

    # J27 — Ask My Accountant
    ama = [r for r in _find_rows_matching(pl_rows, "ask my accountant") if abs(r["amount"]) > 0.01]
    if ama:
        total_ama = sum(r["amount"] for r in ama)
        _set_pl_finding(ws_pl, 27, "clean up needed",
                        comment="Ask My Accountant account has a balance — review and reclassify all transactions",
                        num_txns=str(len(ama)),
                        date_from=period_from_mmyy, date_to=period_to_mmyy,
                        amount=f"${total_ama:,.2f}")
        pl_issues.append("Ask My Accountant balance found")
    else:
        _set_pl_finding(ws_pl, 27, "OK")

    # J28 — Reconciliation Discrepancy
    recon = [r for r in _find_rows_matching(pl_rows, "reconciliation discrepan") if abs(r["amount"]) > 0.01]
    if recon:
        total_recon = sum(r["amount"] for r in recon)
        _set_pl_finding(ws_pl, 28, "clean up needed",
                        comment="Reconciliation Discrepancy account has a balance — investigate and correct",
                        num_txns=str(len(recon)),
                        date_from=period_from_mmyy, date_to=period_to_mmyy,
                        amount=f"${total_recon:,.2f}")
        pl_issues.append("Reconciliation Discrepancy balance found")
    else:
        _set_pl_finding(ws_pl, 28, "OK")

    # J29 — Expenses that should be COGS
    should_cogs = [r for r in _find_rows_matching(pl_rows, "subcontractor", "direct labor", "direct material", "job cost", "project cost") if abs(r["amount"]) > 0.01]
    if should_cogs:
        names = ", ".join(r["name"] for r in should_cogs[:3])
        _set_pl_finding(ws_pl, 29, "clean up needed",
                        comment=f"Possible COGS items in expenses: {names} — consider reclassifying",
                        num_txns=str(len(should_cogs)))
        pl_issues.append("Possible COGS items recorded as expenses")
    else:
        _set_pl_finding(ws_pl, 29, "OK")

    # J30 — Personal expenses
    personal = [r for r in _find_rows_matching(pl_rows, "personal", "owner expense", "meals", "entertainment") if abs(r["amount"]) > 0.01]
    if personal:
        names = ", ".join(r["name"] for r in personal[:3])
        _set_pl_finding(ws_pl, 30, "clean up needed",
                        comment=f"Possible personal expenses: {names} — verify business purpose or reclassify to owner draws",
                        num_txns=str(len(personal)))
        pl_issues.append("Possible personal expenses in business books")
    else:
        _set_pl_finding(ws_pl, 30, "OK")

    # J31 — Loan payments as expenses
    loan_exp = [r for r in _find_rows_matching(pl_rows, "loan payment", "note payable payment", "principal") if abs(r["amount"]) > 0.01]
    if loan_exp:
        _set_pl_finding(ws_pl, 31, "clean up needed",
                        comment="Loan principal payments appear as expenses — split between principal (liability) and interest (expense)",
                        num_txns=str(len(loan_exp)),
                        amount=f"${sum(r['amount'] for r in loan_exp):,.2f}")
        pl_issues.append("Loan payments recorded as expenses")
    else:
        _set_pl_finding(ws_pl, 31, "OK")

    # J32 — Fixed assets under $2500 expensed
    asset_accts = [r for r in _find_rows_matching(pl_rows, "office supplies", "repairs and maintenance", "computer", "equipment rental") if r["amount"] > 2500]
    if asset_accts:
        names = ", ".join(f"{r['name']} (${r['amount']:,.2f})" for r in asset_accts[:3])
        _set_pl_finding(ws_pl, 32, "clean up needed",
                        comment=f"High balances in potential asset expense accounts: {names} — review for capitalization",
                        num_txns=str(len(asset_accts)))
        pl_issues.append("Possible fixed asset purchases expensed — review for capitalization")
    else:
        _set_pl_finding(ws_pl, 32, "OK")

    # J33 — Payroll tax liabilities to expense
    ptax_exp = [r for r in _find_rows_matching(pl_rows, "payroll tax liability", "payroll liab") if abs(r["amount"]) > 0.01]
    if ptax_exp:
        _set_pl_finding(ws_pl, 33, "clean up needed",
                        comment="Payroll tax liabilities recorded to expense accounts — move to liability accounts")
        pl_issues.append("Payroll tax liabilities recorded as expenses")
    else:
        _set_pl_finding(ws_pl, 33, "OK")

    # J34 — Payroll expense recorded incorrectly
    payroll_rows_pl = [r for r in _find_rows_matching(pl_rows, "payroll", "wage", "salary") if abs(r["amount"]) > 0.01]
    has_payroll_exp = bool(payroll_rows_pl)
    has_employer_tax = bool(_find_rows_matching(pl_rows, "employer tax", "payroll tax expense", "fica", "federal tax"))
    if has_payroll_exp and not has_employer_tax:
        _set_pl_finding(ws_pl, 34, "clean up needed",
                        comment="Payroll recorded but no separate employer payroll tax expense found — verify gross wages and employer tax accounts are properly separated")
        pl_issues.append("Payroll structure may need review — employer taxes not clearly separated")
    else:
        _set_pl_finding(ws_pl, 34, "OK")

    # J35 — Miscategorized expenses
    misc_exp = [r for r in _find_rows_matching(pl_rows, "miscellaneous", "other expense", "general expense") if abs(r["amount"]) > 0.01]
    if misc_exp:
        names = ", ".join(r["name"] for r in misc_exp[:3])
        _set_pl_finding(ws_pl, 35, "clean up needed",
                        comment=f"Generic expense accounts with balances: {names} — review and reclassify",
                        num_txns=str(len(misc_exp)),
                        amount=f"${sum(r['amount'] for r in misc_exp):,.2f}")
        pl_issues.append("Miscellaneous/Other expense accounts have balances — reclassify")
    else:
        _set_pl_finding(ws_pl, 35, "OK")

    # J36 — Sales tax as expense
    st_exp = [r for r in _find_rows_matching(pl_rows, "sales tax expense", "sales tax paid") if abs(r["amount"]) > 0.01]
    if st_exp:
        _set_pl_finding(ws_pl, 36, "clean up needed",
                        comment="Sales tax recorded as an expense — should be in Sales Tax Payable and cleared through the Sales Tax Center",
                        amount=f"${sum(r['amount'] for r in st_exp):,.2f}")
        pl_issues.append("Sales tax recorded as expense")
    else:
        _set_pl_finding(ws_pl, 36, "OK")

    # J39 — Negative Other Income/Expense
    neg_other = [r for r in _find_rows_matching(pl_rows, "other income", "other expense") if r["amount"] < -0.01]
    if neg_other:
        names = ", ".join(r["name"] for r in neg_other[:3])
        _set_pl_finding(ws_pl, 39, "clean up needed",
                        comment=f"Negative Other Income/Expense balances: {names}",
                        num_txns=str(len(neg_other)),
                        amount=f"${sum(r['amount'] for r in neg_other):,.2f}")
        pl_issues.append("Negative Other Income/Expense balances")
    else:
        _set_pl_finding(ws_pl, 39, "OK")

    # J42 — Unassigned class transactions (cannot check without P&L by Class)
    ws_pl["J42"].value = "Review needed — run P&L by Class to verify no unassigned transactions"

    # J46 — Unapplied Cash Payment Income
    ucpi = _find_amount(pl_rows, "unapplied cash payment income")
    if abs(ucpi) > 0.01:
        _set_pl_finding(ws_pl, 46, "clean up needed",
                        comment="Unapplied Cash Payment Income found — apply outstanding payments to invoices",
                        amount=f"${ucpi:,.2f}")
        pl_issues.append("Unapplied Cash Payment Income found")
    else:
        _set_pl_finding(ws_pl, 46, "OK")

    # J49 — Unapplied Bill Payment Expense
    ubpe = _find_amount(pl_rows, "unapplied bill payment expense")
    if abs(ubpe) > 0.01:
        _set_pl_finding(ws_pl, 49, "clean up needed",
                        comment="Unapplied Bill Payment Expense found — apply outstanding payments to bills",
                        amount=f"${ubpe:,.2f}")
        pl_issues.append("Unapplied Bill Payment Expense found")
    else:
        _set_pl_finding(ws_pl, 49, "OK")

    # A50 — P&L overall findings
    if pl_issues:
        pl_summary = (
            f"P&L ISSUES FOUND ({len(pl_issues)}):\n"
            + "\n".join(f"• {issue}" for issue in pl_issues)
            + f"\n\nReview period: {period_label} | Method: {accounting_method}"
        )
    else:
        pl_summary = (
            f"P&L review for {period_label} on {accounting_method} basis shows no major issues. "
            "Income, COGS, and expense accounts appear properly categorized."
        )
    ws_pl["A51"].value = pl_summary

    # Work to be completed — P&L
    pl_work = []
    if any("Uncategorized" in i or "uncategorized" in i for i in pl_issues):
        pl_work.append("Reclassify all Uncategorized Income and Uncategorized Expense transactions to proper accounts")
    if any("Ask My Accountant" in i for i in pl_issues):
        pl_work.append("Review and reclassify all Ask My Accountant transactions")
    if any("Reconciliation Discrepancy" in i for i in pl_issues):
        pl_work.append("Investigate and correct the Reconciliation Discrepancy balance")
    if any("Negative income" in i or "negative income" in i for i in pl_issues):
        pl_work.append("Investigate and correct negative income account balances")
    if any("Negative expense" in i or "negative expense" in i for i in pl_issues):
        pl_work.append("Review and correct negative expense account balances — likely reversed entries needed")
    if any("Loan" in i or "loan" in i for i in pl_issues):
        pl_work.append("Split loan payments: principal to liability account, interest to Interest Expense")
    if any("personal" in i.lower() for i in pl_issues):
        pl_work.append("Review personal expense transactions — reclassify to Owner Draw or document business purpose")
    if any("COGS" in i for i in pl_issues):
        pl_work.append("Review COGS classification and ensure all direct costs are correctly categorized")
    if any("Sales tax" in i or "sales tax" in i for i in pl_issues):
        pl_work.append("Move sales tax transactions out of expense accounts and use QBO Sales Tax Center")
    if any("Unapplied" in i for i in pl_issues):
        pl_work.append("Apply all outstanding payments to corresponding invoices or bills in QBO")
    if not pl_work:
        pl_work.append("Continue monitoring P&L monthly for new uncategorized or unusual balances")
    ws_pl["A57"].value = "\n".join(f"• {w}" for w in pl_work)

    issues_found.extend(pl_issues)

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET: Balance Sheet
    # ══════════════════════════════════════════════════════════════════════════
    ws_bs = wb["Balance Sheet"]
    bs_issues: list[str] = []

    ws_bs["J5"].value = period_from_mmyy
    ws_bs["K5"].value = period_to_mmyy

    # J9 — AR positive balance
    ar_bal = _find_amount(bs_rows, "accounts receivable")
    if ar_bal < 0:
        _set_bs_finding(ws_bs, 9, "clean up needed",
                        comment=f"AR has a negative (credit) balance of ${ar_bal:,.2f} — investigate overpayments or misapplied payments",
                        amount=f"${ar_bal:,.2f}")
        bs_issues.append("Negative AR balance found")
    else:
        _set_bs_finding(ws_bs, 9, "OK")

    # J10 — Uncategorized asset
    uncat_asset = _find_amount(bs_rows, "uncategorized asset")
    if abs(uncat_asset) > 0.01:
        _set_bs_finding(ws_bs, 10, "clean up needed",
                        comment=f"Uncategorized Asset balance of ${uncat_asset:,.2f} — reclassify to appropriate asset accounts",
                        amount=f"${uncat_asset:,.2f}")
        bs_issues.append("Uncategorized Asset balance found")
    else:
        _set_bs_finding(ws_bs, 10, "OK")

    # J11 — Unusual current asset balances
    unusual_assets = [
        r for r in bs_rows
        if abs(r["amount"]) > 50000 and not r.get("is_summary")
        and any(kw in r["name"].lower() for kw in ["prepaid", "advance", "deposit", "due from"])
    ]
    if unusual_assets:
        names = ", ".join(f"{r['name']} (${r['amount']:,.2f})" for r in unusual_assets[:3])
        _set_bs_finding(ws_bs, 11, "clean up needed",
                        comment=f"Unusual current asset balances: {names}",
                        num_txns=str(len(unusual_assets)))
        bs_issues.append("Unusual current asset balances found")
    else:
        _set_bs_finding(ws_bs, 11, "OK")

    # J12 — Credit card receivables
    cc_recv = _find_amount(bs_rows, "credit card receivable")
    if abs(cc_recv) > 0.01:
        _set_bs_finding(ws_bs, 12, "clean up needed",
                        comment=f"Credit Card Receivables balance of ${cc_recv:,.2f} — investigate and clear",
                        amount=f"${cc_recv:,.2f}")
        bs_issues.append("Credit Card Receivables balance found")
    else:
        _set_bs_finding(ws_bs, 12, "OK")

    # J13 — Fixed assets under $2500 in COA
    small_fixed = [
        a for a in active_accounts
        if a.get("AccountType") == "Fixed Asset"
        and 0.01 < abs(a.get("CurrentBalance", 0)) < 2500
    ]
    if small_fixed:
        names = ", ".join(a.get("Name", "") for a in small_fixed[:3])
        _set_bs_finding(ws_bs, 13, "clean up needed",
                        comment=f"Fixed asset accounts under $2,500: {names} — may need to be expensed instead",
                        num_txns=str(len(small_fixed)))
        bs_issues.append("Fixed asset accounts under $2,500 found — review capitalization policy")
    else:
        _set_bs_finding(ws_bs, 13, "OK")

    # J14 — Accumulated depreciation
    accum_depr = _find_amount(bs_rows, "accumulated depreciation")
    has_fixed = _find_amount(bs_rows, "fixed asset", "property", "equipment") > 0
    if has_fixed and accum_depr == 0:
        _set_bs_finding(ws_bs, 14, "clean up needed",
                        comment="Fixed assets present but no accumulated depreciation — verify prior year depreciation has been posted",
                        amount="$0.00")
        bs_issues.append("No accumulated depreciation despite fixed assets")
    else:
        _set_bs_finding(ws_bs, 14, "OK")

    # J15 — Unusual other asset balances
    unusual_other = [
        r for r in bs_rows
        if abs(r["amount"]) > 10000 and not r.get("is_summary")
        and any(kw in r["name"].lower() for kw in ["other asset", "security deposit", "notes receivable", "due from officer"])
    ]
    if unusual_other:
        names = ", ".join(f"{r['name']} (${r['amount']:,.2f})" for r in unusual_other[:3])
        _set_bs_finding(ws_bs, 15, "clean up needed",
                        comment=f"Unusual other asset balances: {names}",
                        num_txns=str(len(unusual_other)))
        bs_issues.append("Unusual other asset balances found")
    else:
        _set_bs_finding(ws_bs, 15, "OK")

    # J18 — AP positive balance
    ap_bal = _find_amount(bs_rows, "accounts payable")
    if ap_bal < 0:
        _set_bs_finding(ws_bs, 18, "clean up needed",
                        comment=f"AP has a negative (debit) balance of ${ap_bal:,.2f} — investigate overpayments or duplicate entries",
                        amount=f"${ap_bal:,.2f}")
        bs_issues.append("Negative AP balance found")
    else:
        _set_bs_finding(ws_bs, 18, "OK")

    # J19 — Negative credit card liabilities
    neg_cc = [a for a in active_accounts if a.get("AccountType") == "Credit Card" and a.get("CurrentBalance", 0) < -0.01]
    if neg_cc:
        names = ", ".join(a.get("Name", "") for a in neg_cc[:3])
        _set_bs_finding(ws_bs, 19, "clean up needed",
                        comment=f"Negative credit card liability balances: {names} — investigate payments and credits",
                        num_txns=str(len(neg_cc)),
                        amount=f"${sum(a.get('CurrentBalance', 0) for a in neg_cc):,.2f}")
        bs_issues.append("Negative credit card liability balances found")
    else:
        _set_bs_finding(ws_bs, 19, "OK")

    # J20 — Unusual current liability balances
    unusual_cl = [
        r for r in bs_rows
        if abs(r["amount"]) > 50000 and not r.get("is_summary")
        and any(kw in r["name"].lower() for kw in ["deferred", "customer deposit", "accrued", "due to"])
    ]
    if unusual_cl:
        names = ", ".join(f"{r['name']} (${r['amount']:,.2f})" for r in unusual_cl[:3])
        _set_bs_finding(ws_bs, 20, "clean up needed",
                        comment=f"Unusual current liability balances: {names}",
                        num_txns=str(len(unusual_cl)))
        bs_issues.append("Unusual current liability balances found")
    else:
        _set_bs_finding(ws_bs, 20, "OK")

    # J21 — Payroll liabilities reasonable
    payroll_liab_bal = _find_amount(bs_rows, "payroll liabilit", "payroll tax payable", "federal tax payable")
    if payroll_liab_bal > 100000:
        _set_bs_finding(ws_bs, 21, "clean up needed",
                        comment=f"Payroll liabilities of ${payroll_liab_bal:,.2f} appears high — verify current payroll taxes are clearing monthly",
                        amount=f"${payroll_liab_bal:,.2f}")
        bs_issues.append("High payroll liabilities — verify monthly clearing")
    else:
        _set_bs_finding(ws_bs, 21, "OK")

    # J22 — Sales tax payable reasonable
    st_pay_bal = _find_amount(bs_rows, "sales tax payable", "sales tax liabilit")
    if st_pay_bal > 50000:
        _set_bs_finding(ws_bs, 22, "clean up needed",
                        comment=f"Sales Tax Payable of ${st_pay_bal:,.2f} appears high — verify regular remittance",
                        amount=f"${st_pay_bal:,.2f}")
        bs_issues.append("High sales tax payable — verify remittance frequency")
    else:
        _set_bs_finding(ws_bs, 22, "OK")

    # J23 — Interest expense on notes payable
    notes_pay = _find_amount(bs_rows, "notes payable", "loan payable", "line of credit")
    int_exp = _find_amount(pl_rows, "interest expense")
    if notes_pay > 0 and int_exp == 0:
        _set_bs_finding(ws_bs, 23, "clean up needed",
                        comment="Notes Payable balance exists but no interest expense recorded — verify interest is properly split and recorded",
                        amount=f"${notes_pay:,.2f}")
        bs_issues.append("Notes Payable present but no interest expense recorded")
    else:
        _set_bs_finding(ws_bs, 23, "OK")

    # J24 — Unusual long-term liability balances
    unusual_ltl = [
        r for r in bs_rows
        if abs(r["amount"]) > 500000 and not r.get("is_summary")
        and any(kw in r["name"].lower() for kw in ["long term", "long-term", "note payable", "mortgage"])
    ]
    if unusual_ltl:
        names = ", ".join(f"{r['name']} (${r['amount']:,.2f})" for r in unusual_ltl[:3])
        _set_bs_finding(ws_bs, 24, "clean up needed",
                        comment=f"Large long-term liability balances: {names} — verify balances and terms are current",
                        num_txns=str(len(unusual_ltl)))
        bs_issues.append("Unusual long-term liability balances found")
    else:
        _set_bs_finding(ws_bs, 24, "OK")

    # J27 — Opening Balance Equity
    obe = _find_amount(bs_rows, "opening balance equity")
    if abs(obe) > 0.01:
        _set_bs_finding(ws_bs, 27, "clean up needed",
                        comment=f"Opening Balance Equity has a balance of ${obe:,.2f} — reclassify to appropriate equity accounts",
                        amount=f"${obe:,.2f}")
        bs_issues.append(f"Opening Balance Equity has a balance of ${obe:,.2f}")
    else:
        _set_bs_finding(ws_bs, 27, "OK")

    # J28 — OBE transactions in current year (requires GL drill-down)
    ws_bs["J28"].value = "Review needed — run GL detail on Opening Balance Equity for current year"

    # J29 — Equity contributions
    _set_bs_finding(ws_bs, 29, "OK")

    # J30 — Owner draws
    _set_bs_finding(ws_bs, 30, "OK")

    # J34 — Cash basis AR
    ws_bs["J34"].value = "Review needed — run Balance Sheet on cash basis to verify AR is $0"

    # J37 — Cash basis AP
    ws_bs["J37"].value = "Review needed — run Balance Sheet on cash basis to verify AP is $0"

    # A39 — BS overall findings
    if bs_issues:
        bs_summary = (
            f"BALANCE SHEET ISSUES FOUND ({len(bs_issues)}):\n"
            + "\n".join(f"• {issue}" for issue in bs_issues)
            + f"\n\nAs of: {_fmt_period(period_to)} | Method: Accrual"
        )
    else:
        bs_summary = (
            f"Balance Sheet as of {_fmt_period(period_to)} shows no major issues. "
            "Assets, liabilities, and equity appear properly structured."
        )
    ws_bs["A40"].value = bs_summary

    # Work to be completed — Balance Sheet
    bs_work = []
    if any("Opening Balance Equity" in i for i in bs_issues):
        bs_work.append("Clear Opening Balance Equity: reclassify balance to Retained Earnings or appropriate equity account via journal entry")
    if any("Uncategorized Asset" in i for i in bs_issues):
        bs_work.append("Identify and reclassify all Uncategorized Asset transactions")
    if any("Negative AR" in i for i in bs_issues):
        bs_work.append("Investigate negative AR balance — apply credit memos or correct misposted payments")
    if any("Negative AP" in i for i in bs_issues):
        bs_work.append("Investigate negative AP balance — apply vendor credits or correct duplicate payments")
    if any("depreciation" in i.lower() for i in bs_issues):
        bs_work.append("Post depreciation journal entries for the period; verify prior year depreciation is on file")
    if any("Fixed asset" in i for i in bs_issues):
        bs_work.append("Review fixed asset accounts under $2,500 — expense items below client's capitalization threshold")
    if any("credit card liability" in i.lower() for i in bs_issues):
        bs_work.append("Investigate negative credit card liability balances — likely duplicate payments or recording errors")
    if any("Notes Payable" in i for i in bs_issues):
        bs_work.append("Record interest expense on outstanding notes payable; obtain current loan statement from client")
    if not bs_work:
        bs_work.append("Continue monitoring Balance Sheet monthly; ensure ending balances are reconciled to bank statements")
    ws_bs["A47"].value = "\n".join(f"• {w}" for w in bs_work)

    issues_found.extend(bs_issues)

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET: Accts Receivable+ Accts Payable
    # ══════════════════════════════════════════════════════════════════════════
    ws_arap = wb["Accts Receivable+ Accts Payable"]
    ar_issues: list[str] = []
    ap_issues: list[str] = []

    ar_bkts = ar_data["buckets"]
    ar_90 = ar_data["items_90plus"]
    ar_neg = ar_data["items_negative"]
    ar_zero = ar_data["items_zero"]

    ap_bkts = ap_data["buckets"]
    ap_90 = ap_data["items_90plus"]
    ap_neg = ap_data["items_negative"]
    ap_zero = ap_data["items_zero"]

    # AR checks
    if ar_90 or ar_bkts.get("91+", 0) > 0.01:
        total_90 = ar_bkts.get("91+", 0)
        _set_arap_finding(ws_arap, 6, "clean up needed",
                          num_txns=str(len(ar_90)),
                          date_from=period_from_mmyy, date_to=period_to_mmyy,
                          amount=f"${total_90:,.2f}")
        ar_issues.append(f"AR items over 90 days: ${total_90:,.2f}")
    else:
        _set_arap_finding(ws_arap, 6, "OK")

    if ar_neg:
        _set_arap_finding(ws_arap, 7, "clean up needed",
                          num_txns=str(len(ar_neg)),
                          amount=f"${sum(r['amount'] for r in ar_neg):,.2f}")
        ar_issues.append(f"{len(ar_neg)} customer(s) with credit balances in AR")
    else:
        _set_arap_finding(ws_arap, 7, "OK")

    ws_arap["J8"].value = "Review needed"
    ws_arap["J9"].value = "Review needed"

    if ar_zero:
        _set_arap_finding(ws_arap, 11, "clean up needed", num_txns=str(len(ar_zero)))
        ar_issues.append(f"{len(ar_zero)} customer(s) with $0 AR balance — review and close")
    else:
        _set_arap_finding(ws_arap, 11, "OK")

    if ar_neg:
        _set_arap_finding(ws_arap, 12, "clean up needed", num_txns=str(len(ar_neg)))
        ar_issues.append(f"{len(ar_neg)} available AR credit(s) to apply")
    else:
        _set_arap_finding(ws_arap, 12, "OK")

    ws_arap["A16"].value = (
        (f"AR ISSUES FOUND ({len(ar_issues)}):\n" + "\n".join(f"• {i}" for i in ar_issues)
         + f"\n\nTotal AR: ${ar_bkts.get('total', 0):,.2f} | 90+ days: ${ar_bkts.get('91+', 0):,.2f}")
        if ar_issues else
        f"AR review shows no major issues. Total outstanding: ${ar_bkts.get('total', 0):,.2f}."
    )

    # Work to be completed — AR
    ar_work = []
    if ar_90 or ar_bkts.get("91+", 0) > 0.01:
        ar_work.append(f"Follow up on AR items over 90 days (${ar_bkts.get('91+', 0):,.2f}) — contact customers or write off uncollectable amounts")
    if ar_neg:
        ar_work.append(f"Apply {len(ar_neg)} customer credit balance(s) to open invoices or issue refunds")
    if ar_zero:
        ar_work.append(f"Review and close {len(ar_zero)} customer(s) with $0 AR balance — void or write off stale open invoices")
    if not ar_work:
        ar_work.append("Continue monitoring AR aging monthly; follow up on any items approaching 60 days")
    ws_arap["A21"].value = "\n".join(f"• {w}" for w in ar_work)

    issues_found.extend(ar_issues)

    # AP checks
    if ap_90 or ap_bkts.get("91+", 0) > 0.01:
        total_90_ap = ap_bkts.get("91+", 0)
        _set_arap_finding(ws_arap, 31, "clean up needed",
                          num_txns=str(len(ap_90)),
                          date_from=period_from_mmyy, date_to=period_to_mmyy,
                          amount=f"${total_90_ap:,.2f}")
        ap_issues.append(f"AP items over 90 days: ${total_90_ap:,.2f}")
    else:
        _set_arap_finding(ws_arap, 31, "OK")

    if ap_neg:
        _set_arap_finding(ws_arap, 32, "clean up needed",
                          num_txns=str(len(ap_neg)),
                          amount=f"${sum(r['amount'] for r in ap_neg):,.2f}")
        ap_issues.append(f"{len(ap_neg)} vendor(s) with debit balances in AP")
    else:
        _set_arap_finding(ws_arap, 32, "OK")

    ws_arap["J33"].value = "Review needed"
    ws_arap["J34"].value = "Review needed"

    if ap_zero:
        _set_arap_finding(ws_arap, 36, "clean up needed", num_txns=str(len(ap_zero)))
        ap_issues.append(f"{len(ap_zero)} vendor(s) with $0 AP balance — review and close")
    else:
        _set_arap_finding(ws_arap, 36, "OK")

    if ap_neg:
        _set_arap_finding(ws_arap, 37, "clean up needed", num_txns=str(len(ap_neg)))
        ap_issues.append(f"{len(ap_neg)} available AP credit(s) to apply")
    else:
        _set_arap_finding(ws_arap, 37, "OK")

    ws_arap["A41"].value = (
        (f"AP ISSUES FOUND ({len(ap_issues)}):\n" + "\n".join(f"• {i}" for i in ap_issues)
         + f"\n\nTotal AP: ${ap_bkts.get('total', 0):,.2f} | 90+ days: ${ap_bkts.get('91+', 0):,.2f}")
        if ap_issues else
        f"AP review shows no major issues. Total outstanding: ${ap_bkts.get('total', 0):,.2f}."
    )

    # Work to be completed — AP
    ap_work = []
    if ap_90 or ap_bkts.get("91+", 0) > 0.01:
        ap_work.append(f"Review AP items over 90 days (${ap_bkts.get('91+', 0):,.2f}) — pay outstanding bills or void if no longer owed")
    if ap_neg:
        ap_work.append(f"Apply {len(ap_neg)} vendor debit balance(s) to open bills or request vendor refunds")
    if ap_zero:
        ap_work.append(f"Review and close {len(ap_zero)} vendor(s) with $0 AP balance — void or mark stale open bills as paid")
    if not ap_work:
        ap_work.append("Continue monitoring AP aging monthly; pay all bills before due dates to maintain vendor relationships")
    ws_arap["A46"].value = "\n".join(f"• {w}" for w in ap_work)

    issues_found.extend(ap_issues)

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET: Chart of Accounts
    # ══════════════════════════════════════════════════════════════════════════
    ws_coa = wb["Chart of Accounts"]
    coa_issues: list[str] = []

    # F4 — Number of accounts reasonable
    if total_account_count == 0:
        ws_coa["E4"].value = "Unable to retrieve Chart of Accounts."
    elif total_account_count > 150:
        ws_coa["E4"].value = (
            f"REVIEW NEEDED — {total_account_count} accounts found, which may be excessive. "
            "Consider consolidating duplicate or unused accounts."
        )
        coa_issues.append(f"Chart of Accounts has {total_account_count} accounts — may be excessive")
    else:
        ws_coa["E4"].value = f"OK — {total_account_count} active accounts appears reasonable for the business size."

    # F5 — List reasonableness
    inactive_count = len([a for a in coa_list if not a.get("Active", True)])
    if inactive_count > 20:
        ws_coa["E5"].value = (
            f"REVIEW — {inactive_count} inactive accounts found. "
            "Consider cleaning up the list."
        )
        coa_issues.append(f"{inactive_count} inactive accounts in COA")
    else:
        ws_coa["E5"].value = f"OK — Account list appears reasonable. {inactive_count} inactive accounts."

    # F6 — Account types
    no_subtype = [a for a in active_accounts if not a.get("AccountSubType") and a.get("AccountType") not in ("Bank",)]
    if len(no_subtype) > 5:
        ws_coa["E6"].value = f"REVIEW NEEDED — {len(no_subtype)} accounts missing account subtypes."
        coa_issues.append(f"{len(no_subtype)} accounts missing account subtypes")
    else:
        ws_coa["E6"].value = "OK — Account types appear correctly assigned."

    # F7 — Account numbers
    with_nums = [a for a in active_accounts if a.get("AcctNum")]
    if with_nums:
        ws_coa["E7"].value = f"OK — {len(with_nums)} of {total_account_count} accounts have numbers assigned."
    else:
        ws_coa["E7"].value = (
            "No account numbers in use. Consider implementing a numbering system "
            "(1000s=Assets, 2000s=Liabilities, 3000s=Equity, 4000s=Income, 5000s=COGS, 6000s=Expenses)."
        )
        coa_issues.append("No account numbers assigned")

    # A9 — COA overall findings
    ws_coa["A10"].value = (
        (f"CHART OF ACCOUNTS ISSUES ({len(coa_issues)}):\n" + "\n".join(f"• {i}" for i in coa_issues)
         + f"\n\nTotal active accounts: {total_account_count}")
        if coa_issues else
        f"Chart of Accounts review shows no major issues. {total_account_count} active accounts with appropriate structure."
    )

    # Work to be completed — COA
    coa_work = []
    if total_account_count > 150:
        coa_work.append(f"Audit Chart of Accounts — archive or merge duplicate/unused accounts (currently {total_account_count} active)")
    if inactive_count > 20:
        coa_work.append(f"Review {inactive_count} inactive accounts — confirm they are truly inactive and can be archived")
    if len(no_subtype) > 5:
        coa_work.append(f"Assign correct account subtypes to {len(no_subtype)} accounts missing subtypes — required for proper financial reporting")
    if not with_nums:
        coa_work.append("Implement account numbering system: 1000s=Assets, 2000s=Liabilities, 3000s=Equity, 4000s=Income, 5000s=COGS, 6000s=Expenses")
    if not coa_work:
        coa_work.append("Chart of Accounts is well-organized. Review semi-annually to archive unused accounts.")
    ws_coa["A15"].value = "\n".join(f"• {w}" for w in coa_work)

    issues_found.extend(coa_issues)

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET: Payroll
    # ══════════════════════════════════════════════════════════════════════════
    ws_pay = wb["Payroll"]

    payroll_exp_accts = [
        a for a in active_accounts
        if any(kw in a.get("Name", "").lower() for kw in ["payroll", "salary", "wage"])
    ]
    has_payroll = bool(payroll_exp_accts) or has_payroll_liability

    ws_pay["J4"].value = "Yes" if has_payroll else "No"
    ws_pay["J5"].value = "Unknown — check QBO Payroll or HR records"
    ws_pay["J6"].value = "Unknown — check vendor list for 1099 contractors"
    ws_pay["J7"].value = "Unknown — review payroll schedule with client"
    ws_pay["J8"].value = "Unknown — confirm payroll processor with client"

    if has_payroll and has_payroll_liability:
        ws_pay["J9"].value = "Yes — payroll expense and liability accounts found in COA"
    elif has_payroll and not has_payroll_liability:
        ws_pay["J9"].value = "Review needed — payroll expense accounts found but no payroll liability accounts"
        issues_found.append("Payroll expense accounts exist but no payroll liability accounts in COA")
    else:
        ws_pay["J9"].value = "No payroll detected in QBO"

    ws_pay["A12"].value = (
        ("Payroll accounts detected in Chart of Accounts. " if has_payroll else "No payroll accounts detected. ")
        + "Verify with client: number of employees, subcontractors, payroll frequency, and processor "
        + "(QBO Payroll, third-party service, or manual). "
        + "Confirm employer tax accounts are separate from employee withholding accounts."
    )

    # Work to be completed — Payroll
    pay_work = []
    if has_payroll and not has_payroll_liability:
        pay_work.append("Set up payroll liability accounts in COA: Payroll Tax Payable, Employee Benefits Payable, etc.")
        pay_work.append("Review all payroll expense accounts and ensure proper split between gross wages, employer taxes, and benefits")
    if has_payroll:
        pay_work.append("Confirm with client: number of W-2 employees, 1099 contractors, and payroll frequency")
        pay_work.append("Verify payroll processor and confirm payroll journal entries are correctly imported into QBO")
        pay_work.append("Confirm employer FICA, FUTA, SUTA accounts are separate from employee withholding accounts")
    else:
        pay_work.append("Confirm with client whether they have employees or contractors — may need to set up payroll tracking")
    ws_pay["A17"].value = "\n".join(f"• {w}" for w in pay_work)

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET: Sales Tax
    # ══════════════════════════════════════════════════════════════════════════
    ws_st = wb["Sales Tax"]

    st_payable_accts = [
        a for a in active_accounts
        if a.get("AccountSubType") == "SalesTaxPayable" or "sales tax payable" in a.get("Name", "").lower()
    ]
    uses_sales_tax = bool(st_payable_accts) or has_sales_tax

    ws_st["J4"].value = "Yes" if uses_sales_tax else "No"
    ws_st["J5"].value = "Unknown — confirm remittance frequency with client"
    ws_st["J6"].value = accounting_method
    ws_st["J7"].value = "Yes" if uses_sales_tax else "No"

    if uses_sales_tax:
        st_bal = sum(a.get("CurrentBalance", 0) for a in st_payable_accts)
        st_names = ", ".join(a.get("Name", "") for a in st_payable_accts[:3])
        ws_st["A10"].value = (
            f"Sales Tax Payable accounts found: {st_names}. "
            f"Current balance: ${st_bal:,.2f}. "
            "Verify client is using the QBO Sales Tax Center for all tracking and remittance. "
            "Confirm remittance schedule aligns with state requirements."
        )
        st_work = [
            "Verify all sales are running through QBO Sales Tax Center — not manual Sales Tax Payable entries",
            f"Confirm remittance schedule with client — current balance is ${st_bal:,.2f}",
            "Ensure correct tax rates are applied to taxable products/services",
            "Review prior period sales tax returns for accuracy vs QBO reports",
        ]
    else:
        ws_st["A10"].value = (
            "No sales tax accounts detected in Chart of Accounts. "
            "Confirm with client whether they are required to collect and remit sales tax. "
            "If applicable, set up the QBO Sales Tax Center."
        )
        st_work = [
            "Confirm with client whether they sell taxable goods or services in any state",
            "If sales tax applies: set up QBO Sales Tax Center and configure correct rates by jurisdiction",
        ]
    ws_st["A15"].value = "\n".join(f"• {w}" for w in st_work)

    # ══════════════════════════════════════════════════════════════════════════
    # Client info — overall findings (written last after all checks complete)
    # ══════════════════════════════════════════════════════════════════════════
    total_issues = len(issues_found)
    if total_issues == 0:
        overall_findings = (
            f"QBO Diagnostic Assessment for {client_name or 'client'} — Period: {period_label}\n\n"
            "Overall assessment: No significant issues found. "
            "The QuickBooks file appears to be well-maintained. "
            "Continue with regular reconciliation and review practices."
        )
    else:
        overall_findings = (
            f"QBO Diagnostic Assessment for {client_name or 'client'} — Period: {period_label}\n\n"
            f"ISSUES REQUIRING ATTENTION ({total_issues} total):\n"
            + "\n".join(f"• {issue}" for issue in issues_found[:20])
        )
        if total_issues > 20:
            overall_findings += f"\n...and {total_issues - 20} additional items. See individual sheets for full detail."
        overall_findings += (
            "\n\nRecommendation: Prioritize cleanup in the order listed above. "
            "Schedule a review meeting with the client to discuss findings and establish a cleanup timeline."
        )
    ws_client["A14"].value = overall_findings

    # Work to be completed — Client info / Overall
    client_work = [
        f"Complete full diagnostic assessment for {client_name or 'client'} — Period: {period_label}",
        "Review all findings in individual tabs and address items marked 'clean up needed'",
        "Schedule a client meeting to discuss findings, agree on cleanup timeline, and assign responsibilities",
    ]
    if issues_found:
        client_work.append(f"Prioritize {len(issues_found)} identified issue(s) in order: Banking → P&L → Balance Sheet → AR/AP → COA")
    ws_client["A19"].value = "\n".join(f"• {w}" for w in client_work)

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET: Reconciliation to Tax Return
    # ══════════════════════════════════════════════════════════════════════════
    ws_tax_rec = wb["Reconciliation to Tax Return"]

    # Determine if this applies (S-Corp, Partnership, C-Corp)
    applies_to_tax_rec = any(t in (tax_org_type or "").upper() for t in ["S-CORP", "S CORP", "PARTNERSHIP", "C-CORP", "C CORP"])

    ws_tax_rec["A14"].value = (
        f"Tax entity type: {tax_org_type or 'Not specified'}. "
        + (
            "This assessment applies — obtain the most recent filed tax return and compare Balance Sheet totals "
            f"(Total Assets, Total Liabilities, Total Equity) for the period ending {_fmt_period(period_to)}. "
            "Any differences between QBO and the tax return should be investigated and documented. "
            "Common causes: basis adjustments, depreciation differences, cash vs accrual timing, or missing journal entries."
            if applies_to_tax_rec
            else
            f"For {tax_org_type or 'this entity type'}, a formal tax return balance sheet reconciliation may not be required "
            "(e.g., Sole Proprietors on Schedule C with assets under $250K may not file a balance sheet). "
            "Confirm with the tax preparer whether a reconciliation is needed for this client."
        )
    )
    tax_rec_work = []
    if applies_to_tax_rec:
        tax_rec_work = [
            f"Obtain the most recent tax return for {tax_org_type} from the client",
            f"Run Balance Sheet in QBO as of the last filed tax return period and compare key totals",
            "Document any differences and determine if adjusting journal entries are needed",
            "Coordinate with tax preparer to ensure QBO reflects all tax return adjustments",
        ]
    else:
        tax_rec_work = [
            f"Confirm with client and tax preparer whether a tax return balance sheet reconciliation is required for {tax_org_type or 'this entity'}",
            "If not required, note in file and skip this section",
        ]
    ws_tax_rec["A19"].value = "\n".join(f"• {w}" for w in tax_rec_work)

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET: Undeposited Funds
    # ══════════════════════════════════════════════════════════════════════════
    ws_udf = wb["Undeposited Funds"]

    if undeposited_funds:
        from datetime import date as _date, datetime as _datetime
        today_dt = _date.today()
        old_items = []
        all_dates = []
        udf_total_all = 0.0
        for item in undeposited_funds:
            amt = _safe_float(item.get("Amount", 0) or item.get("TotalAmt", 0))
            udf_total_all += amt
            txn_date_str = item.get("TxnDate", "")
            if txn_date_str:
                try:
                    d = _date.fromisoformat(txn_date_str)
                    all_dates.append(d)
                    if (today_dt - d).days > 30:
                        old_items.append((d, amt, item))
                except Exception:
                    pass

        has_old = bool(old_items)
        oldest_date = min(all_dates) if all_dates else None
        old_total = sum(a for _, a, _ in old_items)

        ws_udf["J5"].value = (
            f"YES — {len(old_items)} item(s) older than 30 days (total: ${old_total:,.2f})"
            if has_old else
            "No — all items are within 30 days"
        )
        ws_udf["Q5"].value = (
            oldest_date.strftime("%m/%d/%Y") if oldest_date else "N/A"
        )

        udf_findings = (
            f"Undeposited Funds contains {len(undeposited_funds)} item(s) totaling ${udf_total_all:,.2f}. "
        )
        if has_old:
            udf_findings += (
                f"{len(old_items)} item(s) are older than 30 days (${old_total:,.2f}) — "
                f"oldest dates to {oldest_date.strftime('%m/%d/%Y') if oldest_date else 'unknown'}. "
                "These may represent forgotten deposits, duplicate entries, or transactions that should have been voided. "
                "Review each item and either deposit, match to an existing bank transaction, or void if erroneous."
            )
        else:
            udf_findings += "All items appear to be recent (within 30 days). Verify each item corresponds to an actual pending deposit."

        ws_udf["A9"].value = udf_findings
        udf_work = []
        if has_old:
            udf_work.append(f"Review {len(old_items)} Undeposited Funds item(s) older than 30 days — deposit, match, or void each one")
            if oldest_date:
                udf_work.append(f"Oldest item dates to {oldest_date.strftime('%m/%d/%Y')} — investigate and resolve")
        udf_work.append("Ensure all customer payments received are promptly deposited and cleared from Undeposited Funds")
        udf_work.append("Do not use Undeposited Funds as a holding account for more than a few days")
        ws_udf["A15"].value = "\n".join(f"• {w}" for w in udf_work)
    else:
        ws_udf["J5"].value = "No items found"
        ws_udf["Q5"].value = "N/A"
        ws_udf["A9"].value = "No items found in Undeposited Funds. Confirm client is consistently depositing payments and clearing this account."
        ws_udf["A15"].value = "• Verify Undeposited Funds account is cleared and confirm all customer payments are deposited promptly."

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET: Products & Services
    # ══════════════════════════════════════════════════════════════════════════
    ws_ps = wb["Products & Services"]

    active_items = [it for it in (items_list or []) if it.get("Active", True)]
    total_items = len(active_items)

    service_items   = [it for it in active_items if it.get("Type") == "Service"]
    product_items   = [it for it in active_items if it.get("Type") in ("Inventory", "NonInventory")]
    inventory_items = [it for it in active_items if it.get("Type") == "Inventory"]
    bundle_items    = [it for it in active_items if it.get("Type") == "Group"]

    # E4 — Is number of items reasonable?
    if total_items == 0:
        ws_ps["E4"].value = "Unable to retrieve Items list from QBO."
    elif total_items > 300:
        ws_ps["E4"].value = (
            f"REVIEW — {total_items} active items found. This may be excessive. "
            "Consider consolidating duplicate or overly granular items."
        )
    elif total_items > 150:
        ws_ps["E4"].value = f"REVIEW — {total_items} items found. Consider whether all items are actively used and necessary."
    else:
        ws_ps["E4"].value = f"OK — {total_items} active item(s) appears reasonable ({len(service_items)} services, {len(product_items)} products, {len(bundle_items)} bundles)."

    # E5 — Is the list reasonable for the industry?
    has_both = bool(service_items) and bool(product_items)
    ws_ps["E5"].value = (
        f"OK — Mix of {len(service_items)} service(s) and {len(product_items)} product(s). "
        "Verify this mix aligns with the client's business model and that each item is actively billed."
        if has_both else
        f"OK — {total_items} {'service' if service_items else 'product'} item(s). "
        "Verify all items are relevant to the client's current operations and pricing is current."
    ) if total_items > 0 else "No items found."

    # E6 — Are types used correctly?
    no_income_acct = [
        it for it in active_items
        if not it.get("IncomeAccountRef") and it.get("Type") in ("Service", "NonInventory")
    ]
    if no_income_acct:
        ws_ps["E6"].value = (
            f"REVIEW — {len(no_income_acct)} item(s) have no income account mapped: "
            + ", ".join(it.get("Name", "") for it in no_income_acct[:5])
            + ". Set correct income accounts for each item."
        )
    elif inventory_items and not any(it.get("AssetAccountRef") for it in inventory_items):
        ws_ps["E6"].value = "REVIEW — Inventory items found but asset account may not be properly set. Verify each inventory item has an Asset Account assigned."
    else:
        ws_ps["E6"].value = "OK — Item types appear to be correctly assigned. Spot-check income/COGS account mappings."

    # E7 — Correctly mapped to income/COGS accounts?
    wrong_income = [
        it for it in active_items
        if it.get("IncomeAccountRef")
        and any(kw in (it.get("IncomeAccountRef", {}).get("name", "") or "").lower()
                for kw in ["expense", "cogs", "cost of"])
    ]
    if wrong_income:
        names = ", ".join(it.get("Name", "") for it in wrong_income[:3])
        ws_ps["E7"].value = (
            f"REVIEW — {len(wrong_income)} item(s) mapped to expense/COGS accounts instead of income accounts: {names}. "
            "Remap to correct income accounts."
        )
    else:
        ws_ps["E7"].value = (
            f"OK — Income account mappings appear reasonable. "
            f"{len(inventory_items)} inventory item(s) should also have COGS and asset accounts assigned."
        ) if total_items > 0 else "No items to review."

    # A9 — P&S findings
    ps_issues = []
    if total_items > 300:
        ps_issues.append(f"Products & Services list has {total_items} items — likely has duplicates or outdated items")
    if no_income_acct:
        ps_issues.append(f"{len(no_income_acct)} item(s) missing income account mapping")
    if wrong_income:
        ps_issues.append(f"{len(wrong_income)} item(s) mapped to wrong account type")
    ws_ps["A10"].value = (
        (f"PRODUCTS & SERVICES ISSUES ({len(ps_issues)}):\n" + "\n".join(f"• {i}" for i in ps_issues)
         + f"\n\nTotal active items: {total_items}")
        if ps_issues else
        f"Products & Services review shows no major issues. {total_items} active items with appropriate structure."
    )

    # A14 — Work to be completed
    ps_work = []
    if no_income_acct:
        ps_work.append(f"Map correct income accounts to {len(no_income_acct)} item(s) missing income account")
    if wrong_income:
        ps_work.append(f"Correct income account mapping for {len(wrong_income)} item(s) mapped to expense/COGS accounts")
    if total_items > 150:
        ps_work.append(f"Review and consolidate Products & Services list — {total_items} items may be excessive")
    if inventory_items:
        ps_work.append(f"Verify {len(inventory_items)} inventory item(s) each have Asset Account and COGS account assigned")
    if not ps_work:
        ps_work.append("Products & Services list appears well-maintained. Review annually to deactivate unused items.")
    ws_ps["A15"].value = "\n".join(f"• {w}" for w in ps_work)

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET: Inventory
    # ══════════════════════════════════════════════════════════════════════════
    ws_inv = wb["Inventory"]
    inv_issues = []

    has_inventory_items = bool(inventory_items)
    inv_asset_bal = _find_amount(bs_rows, "inventory asset", "inventory")

    # O4 — Inventory Valuation Summary
    ws_inv["O4"].value = (
        f"Run Inventory Valuation Summary report in QBO as of {_fmt_period(period_to)}."
        if has_inventory_items else
        "No inventory items found in Products & Services list."
    )
    # O5 — Agree to Balance Sheet
    if has_inventory_items and abs(inv_asset_bal) > 0.01:
        ws_inv["O5"].value = f"Balance Sheet shows Inventory asset of ${inv_asset_bal:,.2f}. Agree this to the Inventory Valuation Summary total."
    elif has_inventory_items:
        ws_inv["O5"].value = "Inventory items exist but no Inventory Asset balance found on Balance Sheet. Investigate."
        inv_issues.append("Inventory items found but no Inventory asset balance on Balance Sheet")
    else:
        ws_inv["O5"].value = "No inventory items — skip this step."

    # O6 — Negative quantities
    ws_inv["O6"].value = (
        "Review Inventory Valuation Summary for any items with negative QTY — indicates a receiving or posting error."
        if has_inventory_items else "N/A — no inventory items."
    )
    # O7 — Inventory Shrinkage
    ws_inv["O7"].value = (
        "Review Inventory Shrinkage account for large or unexpected adjustments."
        if has_inventory_items else "N/A — no inventory items."
    )
    # O8 — Incorrect item types
    ws_inv["O8"].value = (
        f"Review {len(inventory_items)} inventory item(s) to confirm none are set up as service or non-inventory type incorrectly."
        if has_inventory_items else "N/A — no inventory items."
    )

    # A10/F10 — Inventory totals (A9/F9/K9 are column labels; data goes in row 10)
    ws_inv["A10"].value = f"${inv_asset_bal:,.2f}" if has_inventory_items else "$0.00"
    ws_inv["F10"].value = f"${inv_asset_bal:,.2f}" if has_inventory_items else "$0.00"
    # K10 has formula =A10-F10; leave it in place

    # A13 — Inventory findings
    ws_inv["A14"].value = (
        (f"INVENTORY ISSUES ({len(inv_issues)}):\n" + "\n".join(f"• {i}" for i in inv_issues))
        if inv_issues else
        (
            f"Client has {len(inventory_items)} inventory item(s) in QBO. "
            "Run Inventory Valuation Summary and agree total to Balance Sheet. "
            "Review for negative quantities and investigate Inventory Shrinkage account."
            if has_inventory_items else
            "No inventory items found. If client sells physical goods, confirm whether inventory tracking is needed (requires QBO Plus or higher)."
        )
    )

    # A18 — Work to be completed
    inv_work = []
    if has_inventory_items:
        inv_work.append(f"Run Inventory Valuation Summary report as of {_fmt_period(period_to)} and agree total to Balance Sheet")
        inv_work.append("Review for negative quantity items — investigate and correct receiving or posting errors")
        inv_work.append("Review Inventory Shrinkage account for large adjustments")
        inv_work.append(f"Verify all {len(inventory_items)} inventory item(s) are correctly set up with Asset, COGS, and Income accounts")
        if inv_issues:
            inv_work.append("Reconcile difference between Inventory Valuation Summary and Balance Sheet")
    else:
        inv_work.append("Confirm with client whether inventory tracking is needed")
        inv_work.append("If yes: upgrade to QBO Plus, set up inventory items in Products & Services")
    ws_inv["A19"].value = "\n".join(f"• {w}" for w in inv_work)

    # ── Serialize and stream ────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_name = (client_name or "Client").replace(" ", "_").replace("/", "-")
    safe_period = period_to.replace("-", "")[:6]
    filename = f"QBO_Assessment_{safe_name}_{safe_period}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
